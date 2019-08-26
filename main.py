# -*- coding: utf-8 -*-
# @Time    : 2019/8/19 11:00
# @Author  : SDYL
# @Email   : sdyl_1020@163.com
# @File    : main.py
# @Software: PyCharm

import numpy as np
import openpyxl
import matplotlib as mpl
import matplotlib.pyplot as plt


def getXlsx():
    # 获取 工作簿对象
    workbook = openpyxl.load_workbook("Project - venue.xlsx")  # 文件路径
    # 获取工作簿 workbook的所有工作表
    shenames = workbook.sheetnames
    # 获得工作簿的表名后，就可以获得表对象
    # 获取Sheet1表对象
    worksheet_Project = workbook[shenames[0]]
    # 获取Sheet2表对象
    worksheet_Venue = workbook[shenames[1]]
    worksheet_N = workbook[shenames[2]]

    tableName0 = worksheet_Project.title
    tableName1 = worksheet_Venue.title
    tableName2 = worksheet_N.title
    print(tableName0)
    print(tableName1)
    print(tableName2)
    # 获取该表相应的行数和列数
    rows0 = worksheet_Project.max_row
    columns0 = worksheet_Project.max_column
    # 获取该表相应的行数和列数
    print(rows0, columns0)
    rows = worksheet_Venue.max_row
    columns = worksheet_Venue.max_column
    print(rows, columns)
    # # 输出行
    # for row in worksheet_Project.rows:
    #     for cell in row:
    #         print(cell.value, end=" ")
    #     print()
    # 输出行
    N = []
    for row in worksheet_N.rows:
        for cell in row:
            if cell.value == None:
                N.append([0])
            else:
                N.append([1])
    print(N)

        # print(cell.value, end=" ")
        # print()
    '''
    # 输出特定的列
    list00 = []
    for cell in list(worksheet_Project.columns)[0]:  # 获取第1列的数据
        list00.append(cell.value)
    print(list00)

    # 输出特定的列
    list01 = []
    for cell in list(worksheet_Venue.columns)[1]:  # 获取第1列的数据
        list01.append(cell.value)
    print(list01)
    '''

# 输入喜爱项目 形成偏好向量
def setProject():
    pass
# 将项目对应的场馆写入矩阵
# 拟定27个项目分布在35个场馆内 27*35 矩阵表示 项目对应的场馆
# 第一行开始依次表示为 羽毛球、乒乓球、。。。跳伞。跳水
# 第一列表示各个场馆分别为 武汉全民健身中心

def venueProject():
    N = np.zeros((27, 35))
    print(N)
    # 通过逐行遍历 选择 出项目对应的场馆
    # 通过找出的场馆填入矩阵
    for item in N:
        pass

# 通过偏好向量映射对应场馆
def calculate():
    pass


venueProject()
# getXlsx()